import customtkinter as ctk
from tkinter import ttk, messagebox, simpledialog, filedialog
import pandas as pd
import os
import sys
import json
from datetime import datetime, timedelta
from tkcalendar import DateEntry
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from PIL import Image, ImageTk, ImageDraw
import winsound
import threading
import time
import shutil
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
import tkinter as tk
import hashlib
import binascii
import pystray
from pystray import MenuItem as item
import keyboard
import webbrowser
import traceback

# Pre-load matplotlib font cache to avoid blocking
plt.rcParams['font.family'] = 'DejaVu Sans'
plt.rcParams['figure.max_open_warning'] = 50

# Set appearance mode and color theme
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

class AutoLockManager:
    """Auto-lock manager for security"""
    def __init__(self, app, lock_callback, inactivity_minutes=10):
        self.app = app
        self.lock_callback = lock_callback
        self.inactivity_minutes = inactivity_minutes
        self.last_activity_time = time.time()
        self.running = True
        self.setup_auto_lock()
        
    def setup_auto_lock(self):
        """Setup auto-lock functionality"""
        # Bind events to track activity
        events = ['<Button>', '<Key>', '<Motion>']
        for event in events:
            self.app.bind(event, self.reset_inactivity_timer)
        
        # Start checking for inactivity
        self.check_inactivity()
        
    def reset_inactivity_timer(self, event=None):
        """Reset inactivity timer"""
        if self.running and not self.app.is_minimized_to_tray:
            self.last_activity_time = time.time()
        
    def check_inactivity(self):
        """Check for inactivity and lock if needed"""
        if not self.running or not hasattr(self, 'winfo_exists') or not self.winfo_exists():
            return
            
        try:
            # Don't lock if app is minimized to tray
            if self.app.is_minimized_to_tray:
                # Check again after 1 minute if still running
                if self.running and hasattr(self, 'winfo_exists') and self.winfo_exists():
                    self.app.after(60000, self.check_inactivity)
                return
                
            current_time = time.time()
            if current_time - self.last_activity_time > self.inactivity_minutes * 60:
                if self.app.winfo_viewable():  # Only lock if window is visible
                    self.lock_callback()
                    self.reset_inactivity_timer()
            
            # Check again after 1 minute if still running
            if self.running and hasattr(self, 'winfo_exists') and self.winfo_exists():
                self.app.after(60000, self.check_inactivity)
        except tk.TclError:
            # Window has been destroyed
            self.stop()
            
    def stop(self):
        """Stop the auto-lock manager"""
        self.running = False

class PasswordRecovery:
    """Password recovery system"""
    def __init__(self, security_dir):
        self.security_dir = security_dir
        self.recovery_file = os.path.join(security_dir, 'recovery.json')
        
    def has_recovery_data(self):
        """Check if recovery data exists"""
        return os.path.exists(self.recovery_file)
        
    def set_recovery_data(self, question, answer):
        """Set recovery question and answer"""
        recovery_data = {
            "question": question,
            "answer": answer.lower()
        }
        
        with open(self.recovery_file, 'w') as f:
            json.dump(recovery_data, f)
            
    def recover_password(self):
        """Recover password using security question"""
        if not self.has_recovery_data():
            messagebox.showerror("Error", "No recovery data found. Please set up recovery first.")
            return None
            
        try:
            with open(self.recovery_file, 'r') as f:
                recovery_data = json.load(f)
                
            question = recovery_data.get("question", "Security question")
            correct_answer = recovery_data.get("answer", "")
            
            # Ask the security question
            answer = simpledialog.askstring("Password Recovery", f"{question}:")
            if not answer:
                return None
                
            # Check if answer matches (case-insensitive)
            if answer.lower() != correct_answer:
                messagebox.showerror("Error", "Incorrect answer to security question.")
                return None
                
            # Get new password
            new_password = simpledialog.askstring("New Password", "Enter new password:", show='*')
            if not new_password:
                return None
                
            confirm = simpledialog.askstring("Confirm Password", "Confirm new password:", show='*')
            if new_password != confirm:
                messagebox.showerror("Error", "Passwords don't match.")
                return None
                
            return new_password
            
        except Exception as e:
            messagebox.showerror("Error", f"Password recovery failed: {str(e)}")
            return None

class PremiumAirdropTracker(ctk.CTk):
    def __init__(self):
        super().__init__()

        # Configure the main window
        self.title("Premium Airdrop Tracker v1.1")
        self.geometry("1400x800")
        self.minsize(1200, 700)

        # Flags to track application state
        self.running = True
        self.is_minimized_to_tray = False
        self.closing_app = False
        self.data_lock = threading.Lock()  # Lock for thread-safe data access

        # Initialize chart variables
        self.status_fig = None
        self.status_ax = None
        self.status_canvas = None
        self.progress_fig = None
        self.progress_ax = None
        self.progress_canvas = None

        # Password protection
        self.get_data_file_path()
        if not self.check_password():
            self.destroy()
            return

        try:
            # Load data with thread-safe protection
            self.df = self.load_data()
            
            # Create backup directory
            self.create_backup_dir()
            
            # Setup auto backup
            self.setup_auto_backup()
            
            # Setup reminder system
            self.reminder_active = True
            self.shown_reminders = set()
            self.snoozed_projects = {}
            self.reminder_thread = None
            self.reminder_stop_event = threading.Event()
            self.setup_reminder_system()

            # Create the UI widgets
            self.create_widgets()

            # Update UI
            self.update_dashboard()
            self.update_table()
            
            # Track currently editing row
            self.editing_index = None
            
            # Reminder notification window
            self.reminder_window = None
            
            # Setup new features
            self.setup_new_features()
            
            # Load snooze data
            self.load_snooze_data()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to initialize application: {str(e)}")
            self.destroy()

    def setup_new_features(self):
        """Setup new features for v1.1"""
        # System tray
        self.tray_icon = None
        self.tray_thread = None
        self.setup_tray_icon()
        
        # Auto-lock
        self.auto_lock = AutoLockManager(self, self.lock_application, inactivity_minutes=10)
        
        # Password recovery
        self.password_recovery = PasswordRecovery(self.security_dir)

    def setup_tray_icon(self):
        """Setup system tray icon"""
        try:
            # Create a simple icon
            image = Image.new('RGB', (64, 64), color='blue')
            draw = ImageDraw.Draw(image)
            draw.rectangle((16, 16, 48, 48), fill='white')
            
            # Create menu
            menu = (
                item('Show', self.show_from_tray, default=True),
                item('Exit', self.quit_from_tray)
            )
            
            # Create icon but don't start it yet
            self.tray_icon = pystray.Icon("airdrop_tracker", image, "Airdrop Tracker", menu)
            
        except Exception as e:
            print(f"Tray icon setup error: {e}")
            self.tray_icon = None

    def start_tray_icon(self):
        """Start tray icon in separate thread"""
        if self.tray_icon and not self.tray_thread:
            self.tray_thread = threading.Thread(target=self._run_tray_icon, daemon=True)
            self.tray_thread.start()

    def _run_tray_icon(self):
        """Run tray icon (called in separate thread)"""
        try:
            self.tray_icon.run()
        except Exception as e:
            print(f"Tray icon error: {e}")

    def show_from_tray(self, icon=None, item=None):
        """Show application from tray"""
        try:
            self.is_minimized_to_tray = False
            self.deiconify()
            self.lift()
            self.focus_force()
            
            # Hide tray icon when showing app
            if self.tray_icon:
                self.tray_icon.visible = False
                
            # Reset inactivity timer when showing from tray
            if hasattr(self, 'auto_lock'):
                self.auto_lock.reset_inactivity_timer()
        except Exception as e:
            print(f"Error showing from tray: {e}")

    def quit_from_tray(self, icon=None, item=None):
        """Quit application from tray"""
        self.closing_app = True
        self.quit_application()

    def minimize_to_tray(self):
        """Minimize application to system tray"""
        if self.tray_icon:
            try:
                self.is_minimized_to_tray = True
                self.withdraw()
                
                # Start tray icon if not already running
                if not self.tray_thread or not self.tray_thread.is_alive():
                    self.start_tray_icon()
                
                # Make tray icon visible
                self.tray_icon.visible = True
                
            except Exception as e:
                print(f"Error minimizing to tray: {e}")
                messagebox.showerror("Error", "Could not minimize to system tray")
        else:
            messagebox.showerror("Error", "System tray not available")

    def quit_application(self):
        """Quit application completely"""
        try:
            print("Quitting application...")
            self.closing_app = True
            self.running = False
            
            # Stop reminder system
            self.stop_reminder_system()
            
            # Stop auto-lock
            if hasattr(self, 'auto_lock'):
                self.auto_lock.stop()
            
            # Save data
            try:
                self.save_data()
                self.save_snooze_data()
            except:
                pass
            
            # Stop tray icon
            if self.tray_icon:
                try:
                    self.tray_icon.stop()
                except:
                    pass
            
            # Close matplotlib figures
            try:
                if hasattr(self, 'status_fig') and self.status_fig and plt.fignum_exists(self.status_fig.number):
                    plt.close(self.status_fig)
                if hasattr(self, 'progress_fig') and self.progress_fig and plt.fignum_exists(self.progress_fig.number):
                    plt.close(self.progress_fig)
            except:
                pass
            
            # Destroy the window
            try:
                self.quit()
                self.destroy()
            except:
                pass
            
        except Exception as e:
            print(f"Error during quit: {e}")
        finally:
            # Force exit
            os._exit(0)

    def lock_application(self):
        """Lock the application"""
        if not self.is_minimized_to_tray:
            self.withdraw()
            if not self.check_password():
                self.quit_application()
            else:
                self.deiconify()
            
    def initiate_password_recovery(self):
        """Initiate password recovery process"""
        if not self.password_recovery.has_recovery_data():
            # First time setup
            question = simpledialog.askstring("Security Question", 
                                            "Set a security question for password recovery:")
            if not question:
                return
                
            answer = simpledialog.askstring("Security Answer", 
                                          f"Answer for '{question}':")
            if not answer:
                return
                
            self.password_recovery.set_recovery_data(question, answer)
            messagebox.showinfo("Success", "Recovery question set successfully!")
        else:
            # Recovery process
            if messagebox.askyesno("Password Recovery", 
                                 "This will reset your password. Continue?"):
                new_password = self.password_recovery.recover_password()
                if new_password:
                    with open(self.password_file, 'w') as f:
                        f.write(self.hash_password(new_password))
                    messagebox.showinfo("Success", "Password has been reset successfully!")

    def check_password(self):
        """Check if password is set and verify it"""
        if os.path.exists(self.password_file):
            # Password is set, ask for it
            password = simpledialog.askstring("Password", "Enter password:", show='*')
            if password is None:
                return False
            
            with open(self.password_file, 'r') as f:
                stored_password = f.read().strip()
                
            # Verify password
            if not self.verify_password(stored_password, password):
                messagebox.showerror("Error", "Incorrect password")
                return False
        else:
            # First run, set password
            password = simpledialog.askstring("Set Password", "Set a new password:", show='*')
            if password is None or password == "":
                messagebox.showerror("Error", "Password is required")
                return False
                
            confirm = simpledialog.askstring("Confirm Password", "Confirm password:", show='*')
            if password != confirm:
                messagebox.showerror("Error", "Passwords don't match")
                return False
                
            # Hash and store password
            with open(self.password_file, 'w') as f:
                f.write(self.hash_password(password))
                
        return True

    def hash_password(self, password):
        """Hash a password for storing"""
        salt = hashlib.sha256(os.urandom(60)).hexdigest().encode('ascii')
        pwdhash = hashlib.pbkdf2_hmac('sha512', password.encode('utf-8'), salt, 100000)
        pwdhash = binascii.hexlify(pwdhash)
        return (salt + pwdhash).decode('ascii')

    def verify_password(self, stored_password, provided_password):
        """Verify a stored password against a provided password"""
        salt = stored_password[:64].encode('ascii')
        stored_hash = stored_password[64:]
        pwdhash = hashlib.pbkdf2_hmac('sha512', provided_password.encode('utf-8'), salt, 100000)
        pwdhash = binascii.hexlify(pwdhash).decode('ascii')
        return pwdhash == stored_hash

    def change_password(self):
        """Change the application password"""
        old_password = simpledialog.askstring("Change Password", "Enter current password:", show='*')
        if old_password is None:
            return
            
        # Verify old password
        with open(self.password_file, 'r') as f:
            stored_password = f.read().strip()
            
        if not self.verify_password(stored_password, old_password):
            messagebox.showerror("Error", "Incorrect current password")
            return
            
        # Get new password
        new_password = simpledialog.askstring("Change Password", "Enter new password:", show='*')
        if new_password is None or new_password == "":
            return
            
        confirm = simpledialog.askstring("Change Password", "Confirm new password:", show='*')
        if new_password != confirm:
            messagebox.showerror("Error", "Passwords don't match")
            return
            
        # Hash and store new password
        with open(self.password_file, 'w') as f:
            f.write(self.hash_password(new_password))
            
        messagebox.showinfo("Success", "Password changed successfully")

    def get_data_file_path(self):
        """Determine where to save the data file"""
        if getattr(sys, 'frozen', False):
            application_path = os.path.dirname(sys.executable)
        else:
            application_path = os.path.dirname(os.path.abspath(__file__))
        
        self.data_dir = os.path.join(application_path, 'data')
        os.makedirs(self.data_dir, exist_ok=True)
        self.data_file_path = os.path.join(self.data_dir, 'airdrop_data.xlsx')
        self.settings_path = os.path.join(self.data_dir, 'settings.json')
        
        # Create secure directory for password
        self.security_dir = os.path.join(self.data_dir, 'security')
        os.makedirs(self.security_dir, exist_ok=True)
        self.password_file = os.path.join(self.security_dir, 'password.txt')
        
        # Create backup of security directory
        self.backup_dir = os.path.join(self.data_dir, 'backups')
        os.makedirs(self.backup_dir, exist_ok=True)
        self.security_backup_dir = os.path.join(self.backup_dir, 'security')
        os.makedirs(self.security_backup_dir, exist_ok=True)
        
        # Snooze data file
        self.snooze_file = os.path.join(self.data_dir, 'snooze_data.json')

    def create_backup_dir(self):
        """Create backup directory"""
        self.backup_dir = os.path.join(self.data_dir, 'backups')
        os.makedirs(self.backup_dir, exist_ok=True)

    def setup_auto_backup(self):
        """Setup automatic daily backup"""
        self.auto_backup = True
        self.last_backup_time = time.time()
        self.backup_interval = 86400  # 24 hours in seconds
        self.run_backup_check()

    def run_backup_check(self):
        """Check if backup is needed"""
        if not self.running or not hasattr(self, 'auto_backup'):
            return
            
        try:
            if self.auto_backup and (time.time() - self.last_backup_time >= self.backup_interval):
                try:
                    self.create_backup()
                    self.last_backup_time = time.time()
                except Exception as e:
                    print(f"Backup error: {e}")
            
            if self.running and not self.closing_app:
                self.after(3600000, self.run_backup_check)  # Check every hour
        except tk.TclError:
            # Window has been destroyed
            pass

    def setup_reminder_system(self):
        """Setup reminder system with proper thread management"""
        self.reminder_stop_event.clear()
        if self.reminder_thread is None or not self.reminder_thread.is_alive():
            self.reminder_thread = threading.Thread(target=self.reminder_check_loop, daemon=True)
            self.reminder_thread.start()

    def stop_reminder_system(self):
        """Stop reminder system properly"""
        self.reminder_active = False
        self.reminder_stop_event.set()
        
        # Wait for thread to finish (with timeout)
        if self.reminder_thread and self.reminder_thread.is_alive():
            self.reminder_thread.join(timeout=2.0)

    def reminder_check_loop(self):
        """Continuous reminder check loop with proper thread safety"""
        while self.running and self.reminder_active and not self.reminder_stop_event.is_set():
            try:
                if hasattr(self, 'df') and not self.df.empty and not self.closing_app:
                    self.check_reminders()
            except Exception as e:
                print(f"Reminder error: {e}")
            
            # Use event wait instead of sleep for better responsiveness
            if self.reminder_stop_event.wait(timeout=60):
                break

    def create_widgets(self):
        """Create all UI widgets"""
        # Create main container
        self.main_container = ctk.CTkFrame(self)
        self.main_container.pack(fill="both", expand=True, padx=10, pady=10)

        # Create sidebar
        self.create_sidebar()
        
        # Create main content area
        self.create_main_content()
        
        # Add version and copyright info
        self.create_footer()

    def create_footer(self):
        """Create footer with version and copyright info"""
        footer_frame = ctk.CTkFrame(self, height=30, fg_color="transparent")
        footer_frame.pack(side="bottom", fill="x", padx=10, pady=5)
        
        # Version info
        version_label = ctk.CTkLabel(footer_frame, text="v1.1", 
                                   font=ctk.CTkFont(size=12))
        version_label.pack(side="left", padx=10)
        
        # Copyright info
        copyright_label = ctk.CTkLabel(footer_frame, 
                                     text="© 2025 NafasMedia Software. All rights reserved.",
                                     font=ctk.CTkFont(size=12))
        copyright_label.pack(side="right", padx=10)

    def create_sidebar(self):
        """Create sidebar with navigation"""
        self.sidebar = ctk.CTkFrame(self.main_container, width=250)
        self.sidebar.pack(side="left", fill="y", padx=(0, 10), pady=10)
        self.sidebar.pack_propagate(False)

        # Logo/Title
        title_label = ctk.CTkLabel(self.sidebar, text="Airdrop Tracker", 
                                 font=ctk.CTkFont(size=20, weight="bold"))
        title_label.pack(pady=20)

        # Navigation buttons
        nav_buttons = [
            ("📊 Dashboard", self.show_dashboard),
            ("📋 Data Manager", self.show_data_manager),
            ("⏰ Reminders", self.show_reminder_settings),
            ("⚙️ Settings", self.show_settings),
            ("ℹ️ About", self.show_about)
        ]

        for text, command in nav_buttons:
            btn = ctk.CTkButton(self.sidebar, text=text, command=command,
                              fg_color="transparent", hover_color=("gray70", "gray30"),
                              anchor="w", text_color=("gray10", "gray90"))
            btn.pack(fill="x", padx=10, pady=5)

        # Theme switcher
        self.theme_var = ctk.StringVar(value="Dark")
        theme_switch = ctk.CTkSwitch(self.sidebar, text="Light Mode", 
                                   variable=self.theme_var, onvalue="Light", offvalue="Dark",
                                   command=self.toggle_theme)
        theme_switch.pack(pady=20, padx=10)

    def create_main_content(self):
        """Create main content area"""
        self.main_content = ctk.CTkFrame(self.main_container)
        self.main_content.pack(side="right", fill="both", expand=True, pady=10)

        # Create tabs
        self.tabview = ctk.CTkTabview(self.main_content)
        self.tabview.pack(fill="both", expand=True)
        
        # Create tabs
        self.dashboard_tab = self.tabview.add("📊 Dashboard")
        self.data_tab = self.tabview.add("📋 Data")
        self.reminder_tab = self.tabview.add("⏰ Reminders")
        self.settings_tab = self.tabview.add("⚙️ Settings")
        self.about_tab = self.tabview.add("ℹ️ About")

        # Setup each tab
        self.setup_dashboard_tab()
        self.setup_data_tab()
        self.setup_reminder_tab()
        self.setup_settings_tab()
        self.setup_about_tab()

    def setup_about_tab(self):
        """Setup about tab"""
        about_frame = ctk.CTkFrame(self.about_tab, fg_color="transparent")
        about_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # App info
        app_name = ctk.CTkLabel(about_frame, text="Premium Airdrop Tracker", 
                               font=ctk.CTkFont(size=24, weight="bold"))
        app_name.pack(pady=10)
        
        version = ctk.CTkLabel(about_frame, text="Apps Version 1.1", 
                              font=ctk.CTkFont(size=16))
        version.pack(pady=5)
        
        developer = ctk.CTkLabel(about_frame, text="Developed by NafasMedia Software", 
                                font=ctk.CTkFont(size=14))
        developer.pack(pady=5)
        
        copyright = ctk.CTkLabel(about_frame, 
                                text="Copyright 2025 NafasMedia Software. All rights reserved.",
                                font=ctk.CTkFont(size=14))
        copyright.pack(pady=20)
        
        # Website button
        website_btn = ctk.CTkButton(about_frame, text="Visit Our Website", 
                                  command=lambda: webbrowser.open("https://at.nafasmedia.eu.org/"))
        website_btn.pack(pady=10, padx=50, fill="x")
        
        # Privacy Policy button
        privacy_btn = ctk.CTkButton(about_frame, text="Privacy Policy", 
                                  command=lambda: webbrowser.open("https://at.nafasmedia.eu.org/p/privacy.html"))
        privacy_btn.pack(pady=10, padx=50, fill="x")
        
        # Terms of Service button
        tos_btn = ctk.CTkButton(about_frame, text="Terms of Service", 
                              command=lambda: webbrowser.open("https://at.nafasmedia.eu.org/terms.html"))
        tos_btn.pack(pady=10, padx=50, fill="x")

    def setup_dashboard_tab(self):
        """Setup dashboard tab"""
        # Stats frame
        stats_frame = ctk.CTkFrame(self.dashboard_tab)
        stats_frame.pack(fill="x", padx=10, pady=10)

        # Configure grid for responsive layout
        stats_frame.grid_columnconfigure((0, 1, 2, 3), weight=1, uniform="stats_cols")
        stats_frame.grid_rowconfigure(0, weight=1)

        # Stats data with corrected colors
        stats_data = [
            ("Total Projects", len(self.df), "#4ECDC4"),
            ("Active Projects", len(self.df[self.df['Status'] == 'Active']), "#45B7D1"),
            ("Completed", len(self.df[self.df['Status'] == 'Completed']), "#96CEB4"),
            ("Total Value", f"Rp {self.df['Estimated Reward'].sum():,.2f}" if not self.df.empty else "Rp 0", "#FECA57")
        ]

        for i, (title, value, color) in enumerate(stats_data):
            stat_frame = ctk.CTkFrame(stats_frame, fg_color=color, corner_radius=10)
            stat_frame.grid(row=0, column=i, padx=5, pady=5, sticky="nsew")
            
            ctk.CTkLabel(stat_frame, text=title, font=ctk.CTkFont(weight="bold")).pack(pady=5)
            ctk.CTkLabel(stat_frame, text=str(value), font=ctk.CTkFont(size=16)).pack(pady=5)

        # Charts frame
        charts_frame = ctk.CTkFrame(self.dashboard_tab)
        charts_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Configure grid for responsive charts
        charts_frame.grid_columnconfigure((0, 1), weight=1, uniform="chart_cols")
        charts_frame.grid_rowconfigure(0, weight=1)

        # Status distribution chart
        status_frame = ctk.CTkFrame(charts_frame)
        status_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
        self.create_status_chart(status_frame)

        # Progress chart
        progress_frame = ctk.CTkFrame(charts_frame)
        progress_frame.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
        self.create_progress_chart(progress_frame)

    def create_status_chart(self, parent):
        """Create status distribution chart with efficient figure management"""
        try:
            # Clear existing widgets
            for widget in parent.winfo_children():
                widget.destroy()
            
            # Create new figure and axis
            fig, ax = plt.subplots(figsize=(5, 4))
            
            if not self.df.empty and 'Status' in self.df.columns:
                status_counts = self.df['Status'].value_counts()
                if not status_counts.empty:
                    # Corrected colors
                    colors = []
                    for status in status_counts.index:
                        if status == 'Active':
                            colors.append('#45B7D1')  # Light blue
                        elif status == 'Completed':
                            colors.append('#96CEB4')  # Light green
                        elif status == 'Monitoring':
                            colors.append('#FFA500')  # Orange
                        elif status == 'Dropped':
                            colors.append('#FF6B6B')  # Red
                        else:
                            colors.append('#CCCCCC')  # Default
                    
                    ax.pie(status_counts.values, labels=status_counts.index, autopct='%1.1f%%', colors=colors)
                    ax.set_title('Status Distribution')
                else:
                    ax.text(0.5, 0.5, 'No data available', ha='center', va='center')
                    ax.set_title('Status Distribution')
            else:
                ax.text(0.5, 0.5, 'No data available', ha='center', va='center')
                ax.set_title('Status Distribution')
            
            # Create canvas
            canvas = FigureCanvasTkAgg(fig, parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True)
            
            # Store references
            self.status_fig = fig
            self.status_ax = ax
            self.status_canvas = canvas
                
        except Exception as e:
            print(f"Chart error: {e}")

    def create_progress_chart(self, parent):
        """Create progress chart with efficient figure management"""
        try:
            # Clear existing widgets
            for widget in parent.winfo_children():
                widget.destroy()
                
            # Create new figure and axis
            fig, ax = plt.subplots(figsize=(5, 4))
            
            if not self.df.empty and 'Progress' in self.df.columns:
                progress_counts = self.df['Progress'].value_counts()
                if not progress_counts.empty:
                    ax.bar(progress_counts.index, progress_counts.values)
                    ax.set_title('Progress Distribution')
                    ax.set_xlabel('Progress')
                    ax.set_ylabel('Count')
                    plt.xticks(rotation=45)
                else:
                    ax.text(0.5, 0.5, 'No data available', ha='center', va='center')
                    ax.set_title('Progress Distribution')
            else:
                ax.text(0.5, 0.5, 'No data available', ha='center', va='center')
                ax.set_title('Progress Distribution')
            
            # Create canvas
            canvas = FigureCanvasTkAgg(fig, parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True)
            
            # Store references
            self.progress_fig = fig
            self.progress_ax = ax
            self.progress_canvas = canvas
                
        except Exception as e:
            print(f"Chart error: {e}")

    def setup_data_tab(self):
        """Setup data management tab"""
        # Input form
        form_frame = ctk.CTkFrame(self.data_tab)
        form_frame.pack(fill="x", padx=10, pady=10)

        # Form fields
        fields = [
            ("Project Name", "entry", ""),
            ("Status", "option", ["Active", "Completed", "Monitoring", "Dropped"]),
            ("Due Date", "date", ""),
            ("Due Time", "time", ""),
            ("Progress", "option", ["0%", "25%", "50%", "75%", "100%"]),
            ("Estimated Reward (Rp)", "entry", ""),
            ("Project Link", "entry", ""),
            ("Notes", "text", ""),
        ]

        self.form_fields = {}
        for i, (label, field_type, options) in enumerate(fields):
            row = i // 2
            col = (i % 2) * 2
            
            ctk.CTkLabel(form_frame, text=label).grid(row=row, column=col, padx=5, pady=5, sticky="w")
            
            if field_type == "entry":
                entry = ctk.CTkEntry(form_frame, width=200)
                entry.grid(row=row, column=col+1, padx=5, pady=5)
                self.form_fields[label] = entry
            elif field_type == "option":
                var = ctk.StringVar(value=options[0] if options else "")
                option = ctk.CTkOptionMenu(form_frame, variable=var, values=options, width=200)
                option.grid(row=row, column=col+1, padx=5, pady=5)
                self.form_fields[label] = var
            elif field_type == "date":
                date_entry = DateEntry(form_frame, width=18, background='darkblue',
                                     foreground='white', borderwidth=2, date_pattern='y-mm-dd')
                date_entry.grid(row=row, column=col+1, padx=5, pady=5)
                self.form_fields[label] = date_entry
            elif field_type == "time":
                # Create time entry with custom validation
                time_frame = ctk.CTkFrame(form_frame, width=200, height=30)
                time_frame.grid(row=row, column=col+1, padx=5, pady=5)
                time_frame.grid_propagate(False)
                
                # Hours dropdown
                hours = [f"{h:02d}" for h in range(0, 24)]
                hour_var = ctk.StringVar(value="09")
                hour_dropdown = ctk.CTkOptionMenu(time_frame, variable=hour_var, values=hours, width=60)
                hour_dropdown.pack(side="left", padx=2)
                
                # Separator
                ctk.CTkLabel(time_frame, text=":").pack(side="left", padx=2)
                
                # Minutes dropdown with all minutes (0-59)
                minutes = [f"{m:02d}" for m in range(0, 60)]
                minute_var = ctk.StringVar(value="00")
                minute_dropdown = ctk.CTkOptionMenu(time_frame, variable=minute_var, values=minutes, width=60)
                minute_dropdown.pack(side="left", padx=2)
                
                # Store both variables
                self.form_fields[label] = (hour_var, minute_var)
            elif field_type == "text":
                # Notes field
                notes_text = ctk.CTkTextbox(form_frame, width=400, height=80)
                notes_text.grid(row=row, column=col+1, columnspan=3, padx=5, pady=5, sticky="ew")
                self.form_fields[label] = notes_text

        # Buttons
        btn_frame = ctk.CTkFrame(form_frame)
        btn_frame.grid(row=4, column=0, columnspan=4, pady=10)

        self.add_btn = ctk.CTkButton(btn_frame, text="Add Project", command=self.add_project)
        self.add_btn.pack(side="left", padx=5)
        
        ctk.CTkButton(btn_frame, text="Clear Form", command=self.clear_form).pack(side="left", padx=5)
        
        # Update button is always visible but disabled when no selection
        self.update_btn = ctk.CTkButton(btn_frame, text="Update Selected", command=self.update_project, state="disabled")
        self.update_btn.pack(side="left", padx=5)
        
        ctk.CTkButton(btn_frame, text="Delete Selected", command=self.delete_project, 
                     fg_color="red", hover_color="darkred").pack(side="left", padx=5)

        # Data table
        table_frame = ctk.CTkFrame(self.data_tab)
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)

        columns = ["ID", "Project", "Status", "Due DateTime", "Progress", "Reward", "Link", "Notes"]
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        
        # Set column widths
        self.tree.column("ID", width=50)
        self.tree.column("Project", width=150)
        self.tree.column("Status", width=100)
        self.tree.column("Due DateTime", width=180)
        self.tree.column("Progress", width=80)
        self.tree.column("Reward", width=120)
        self.tree.column("Link", width=150)
        self.tree.column("Notes", width=200)
        
        for col in columns:
            self.tree.heading(col, text=col)

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Bind selection
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

    def setup_reminder_tab(self):
        """Setup reminder settings tab"""
        reminder_frame = ctk.CTkFrame(self.reminder_tab)
        reminder_frame.pack(fill="both", expand=True, padx=10, pady=10)

        ctk.CTkLabel(reminder_frame, text="Reminder Settings", 
                   font=ctk.CTkFont(size=16, weight="bold")).pack(pady=10)

        # Reminder toggle
        self.reminder_var = ctk.BooleanVar(value=True)
        self.reminder_toggle = ctk.CTkSwitch(reminder_frame, text="Enable Reminders",
                                           variable=self.reminder_var, command=self.toggle_reminders)
        self.reminder_toggle.pack(pady=5)

        # Test reminder button
        ctk.CTkButton(reminder_frame, text="Test Reminder", 
                     command=self.test_reminder).pack(pady=5)

        # Snooze options
        ctk.CTkLabel(reminder_frame, text="Snooze Duration (minutes):").pack(pady=(20, 5))
        
        self.snooze_var = ctk.StringVar(value="30")
        snooze_options = ["15", "30", "60", "120", "240"]
        snooze_dropdown = ctk.CTkOptionMenu(reminder_frame, variable=self.snooze_var, values=snooze_options)
        snooze_dropdown.pack(pady=5)

        # Per-project reminder settings
        ctk.CTkLabel(reminder_frame, text="Per-Project Reminder Settings", 
                   font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(30, 10))
        
        # Frame for project selection
        project_frame = ctk.CTkFrame(reminder_frame)
        project_frame.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(project_frame, text="Select Project:").pack(side="left", padx=5)
        
        self.project_var = ctk.StringVar(value="")
        self.project_dropdown = ctk.CTkOptionMenu(project_frame, variable=self.project_var, 
                                                values=[], command=self.on_project_select)
        self.project_dropdown.pack(side="left", padx=5, fill="x", expand=True)
        
        # Reminder toggle for selected project
        self.project_reminder_var = ctk.BooleanVar(value=True)
        self.project_reminder_toggle = ctk.CTkSwitch(reminder_frame, text="Enable Reminder",
                                                   variable=self.project_reminder_var, command=self.toggle_project_reminder)
        self.project_reminder_toggle.pack(pady=10)
        
        # Update project dropdown
        self.update_project_dropdown()

    def setup_settings_tab(self):
        """Setup settings tab"""
        # Use a scrollable frame for settings to prevent cutoff
        settings_scroll = ctk.CTkScrollableFrame(self.settings_tab, height=600)
        settings_scroll.pack(fill="both", expand=True, padx=10, pady=10)

        ctk.CTkLabel(settings_scroll, text="Application Settings", 
                   font=ctk.CTkFont(size=16, weight="bold")).pack(pady=10)

        # Password settings
        ctk.CTkLabel(settings_scroll, text="Password Settings", 
                   font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(20, 5))
        
        ctk.CTkButton(settings_scroll, text="Change Password", 
                     command=self.change_password).pack(pady=5)
        
        # Password recovery button
        ctk.CTkButton(settings_scroll, text="Password Recovery", 
                     command=self.initiate_password_recovery).pack(pady=5)

        # Backup settings
        ctk.CTkLabel(settings_scroll, text="Backup Settings").pack(pady=(20, 5))
        self.backup_toggle = ctk.CTkSwitch(settings_scroll, text="Auto Backup",
                                         command=self.toggle_auto_backup)
        self.backup_toggle.pack(pady=5)
        self.backup_toggle.select()  # Default to enabled

        # Manual backup button
        ctk.CTkButton(settings_scroll, text="Create Backup Now", 
                     command=self.create_backup).pack(pady=5)

        # Export options
        ctk.CTkLabel(settings_scroll, text="Export Options").pack(pady=(20, 5))
        
        export_frame = ctk.CTkFrame(settings_scroll)
        export_frame.pack(pady=5)
        
        ctk.CTkButton(export_frame, text="Export to Excel", 
                     command=self.export_excel).pack(side="left", padx=5)
        ctk.CTkButton(export_frame, text="Export to CSV", 
                     command=self.export_csv).pack(side="left", padx=5)
        ctk.CTkButton(export_frame, text="Export to PDF", 
                     command=self.export_pdf).pack(side="left", padx=5)
        
        # Import options
        ctk.CTkLabel(settings_scroll, text="Import Options").pack(pady=(20, 5))
        
        import_frame = ctk.CTkFrame(settings_scroll)
        import_frame.pack(pady=5)
        
        ctk.CTkButton(import_frame, text="Import from Excel", 
                     command=self.import_excel).pack(side="left", padx=5)
        ctk.CTkButton(import_frame, text="Import from CSV", 
                     command=self.import_csv).pack(side="left", padx=5)

        # Window options
        ctk.CTkLabel(settings_scroll, text="Window Options", 
                   font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(20, 5))
        
        ctk.CTkButton(settings_scroll, text="Minimize to Tray", 
                     command=self.minimize_to_tray).pack(pady=5)

        # Delete all data button
        ctk.CTkLabel(settings_scroll, text="Danger Zone", 
                   font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(30, 5))
        
        ctk.CTkButton(settings_scroll, text="Delete All Data", 
                     command=self.delete_all_data, fg_color="red", hover_color="darkred").pack(pady=10)

    def add_project(self):
        """Add new project"""
        try:
            # Get form data
            project_name = self.form_fields['Project Name'].get().strip()
            if not project_name:
                messagebox.showwarning("Warning", "Project Name is required!")
                return

            # Get time from time fields
            hour_var, minute_var = self.form_fields['Due Time']
            due_time = f"{hour_var.get()}:{minute_var.get()}"

            project_data = {
                'Project Name': project_name,
                'Status': self.form_fields['Status'].get(),
                'Due Date': self.form_fields['Due Date'].get_date().strftime('%Y-%m-%d'),
                'Due Time': due_time,
                'Progress': self.form_fields['Progress'].get(),
                'Estimated Reward': float(self.form_fields['Estimated Reward (Rp)'].get() or 0),
                'Project Link': self.form_fields['Project Link'].get().strip(),
                'Notes': self.form_fields['Notes'].get("1.0", "end-1c").strip(),
                'Reminder Enabled': True  # Default to enabled
            }

            # Add to DataFrame with thread safety
            with self.data_lock:
                if self.df.empty:
                    self.df = pd.DataFrame([project_data])
                else:
                    self.df = pd.concat([self.df, pd.DataFrame([project_data])], ignore_index=True)

            # Save and update
            self.save_data()
            self.update_table()
            self.update_dashboard()
            self.clear_form()
            self.update_project_dropdown()

            messagebox.showinfo("Success", "Project added successfully!")

        except ValueError as e:
            messagebox.showwarning("Warning", f"Please enter a valid number for Estimated Reward! Error: {str(e)}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add project: {str(e)}")

    def update_project(self):
        """Update existing project"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a project to update!")
            return

        try:
            index = self.tree.index(selected[0])
            
            # Get form data
            project_name = self.form_fields['Project Name'].get().strip()
            if not project_name:
                messagebox.showwarning("Warning", "Project Name is required!")
                return

            # Get time from time fields
            hour_var, minute_var = self.form_fields['Due Time']
            due_time = f"{hour_var.get()}:{minute_var.get()}"

            # Update the DataFrame row with thread safety
            with self.data_lock:
                self.df.at[index, 'Project Name'] = project_name
                self.df.at[index, 'Status'] = self.form_fields['Status'].get()
                self.df.at[index, 'Due Date'] = self.form_fields['Due Date'].get_date().strftime('%Y-%m-%d')
                self.df.at[index, 'Due Time'] = due_time
                self.df.at[index, 'Progress'] = self.form_fields['Progress'].get()
                self.df.at[index, 'Estimated Reward'] = float(self.form_fields['Estimated Reward (Rp)'].get() or 0)
                self.df.at[index, 'Project Link'] = self.form_fields['Project Link'].get().strip()
                self.df.at[index, 'Notes'] = self.form_fields['Notes'].get("1.0", "end-1c").strip()

            # Remove from shown reminders if due date/time changed
            project_id = f"{project_name}_{self.df.at[index, 'Due Date']}_{self.df.at[index, 'Due Time']}"
            if project_id in self.shown_reminders:
                self.shown_reminders.remove(project_id)

            # Save and update
            self.save_data()
            self.update_table()
            self.update_dashboard()
            self.clear_form()
            self.update_project_dropdown()

            messagebox.showinfo("Success", "Project updated successfully!")

        except ValueError as e:
            messagebox.showwarning("Warning", f"Please enter a valid number for Estimated Reward! Error: {str(e)}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update project: {str(e)}")

    def clear_form(self):
        """Clear all form fields"""
        for field_name, field in self.form_fields.items():
            if isinstance(field, ctk.CTkEntry):
                field.delete(0, 'end')
            elif isinstance(field, ctk.StringVar):
                # Set to first option for dropdowns
                if field_name == "Status":
                    field.set("Active")
                elif field_name == "Progress":
                    field.set("0%")
                else:
                    field.set("")
            elif isinstance(field, tuple):  # Time field
                hour_var, minute_var = field
                hour_var.set("09")
                minute_var.set("00")
        
        # Reset notes
        self.form_fields['Notes'].delete("1.0", "end")
        # Reset date to today
        if 'Due Date' in self.form_fields:
            self.form_fields['Due Date'].set_date(datetime.now())
            
        # Disable update button
        self.update_btn.configure(state="disabled")
        self.add_btn.configure(text="Add Project")

    def on_tree_select(self, event):
        """Handle treeview selection"""
        selected = self.tree.selection()
        if selected:
            index = self.tree.index(selected[0])
            self.editing_index = index
            
            # Enable update button
            self.update_btn.configure(state="normal")
            
            # Populate form with selected data
            row = self.df.iloc[index]
            
            self.form_fields['Project Name'].delete(0, 'end')
            self.form_fields['Project Name'].insert(0, row.get('Project Name', ''))
            
            self.form_fields['Status'].set(row.get('Status', 'Active'))
            
            due_date = row.get('Due Date', '')
            if due_date:
                try:
                    if isinstance(due_date, str):
                        due_date = datetime.strptime(due_date, '%Y-%m-%d')
                    self.form_fields['Due Date'].set_date(due_date)
                except:
                    self.form_fields['Due Date'].set_date(datetime.now())
            
            # Set time fields
            due_time = row.get('Due Time', '09:00')
            if due_time and ':' in due_time:
                hours, minutes = due_time.split(':')
                hour_var, minute_var = self.form_fields['Due Time']
                hour_var.set(hours)
                minute_var.set(minutes)
            
            self.form_fields['Progress'].set(row.get('Progress', '0%'))
            
            self.form_fields['Estimated Reward (Rp)'].delete(0, 'end')
            self.form_fields['Estimated Reward (Rp)'].insert(0, str(row.get('Estimated Reward', 0)))
            
            self.form_fields['Project Link'].delete(0, 'end')
            self.form_fields['Project Link'].insert(0, row.get('Project Link', ''))
            
            self.form_fields['Notes'].delete("1.0", "end")
            self.form_fields['Notes'].insert("1.0", row.get('Notes', ''))

    def delete_project(self):
        """Delete selected project"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a project to delete!")
            return

        if messagebox.askyesno("Confirm", "Are you sure you want to delete the selected project?"):
            try:
                # Get selected index
                selected_index = self.tree.index(selected[0])
                project_name = self.df.at[selected_index, 'Project Name']
                due_date = self.df.at[selected_index, 'Due Date']
                due_time = self.df.at[selected_index, 'Due Time']
                
                # Remove from shown reminders
                project_id = f"{project_name}_{due_date}_{due_time}"
                if project_id in self.shown_reminders:
                    self.shown_reminders.remove(project_id)
                
                # Remove from snoozed projects
                if project_id in self.snoozed_projects:
                    del self.snoozed_projects[project_id]
                
                # Delete with thread safety
                with self.data_lock:
                    self.df = self.df.drop(selected_index).reset_index(drop=True)
                
                self.save_data()
                self.update_table()
                self.update_dashboard()
                self.clear_form()
                self.update_project_dropdown()
                messagebox.showinfo("Success", "Project deleted successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete project: {str(e)}")

    def update_table(self):
        """Update data table with combined date and time"""
        # Clear existing data
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Add new data with combined date/time
        for index, row in self.df.iterrows():
            reward_value = row.get('Estimated Reward', 0)
            if pd.notna(reward_value):
                reward_str = f"Rp {reward_value:,.2f}"
            else:
                reward_str = "Rp 0"
                
            # Combine date and time
            due_date = row.get('Due Date', '')
            due_time = row.get('Due Time', '')
            due_datetime = f"{due_date} {due_time}" if due_date and due_time else ''
            
            # Add day name if date is available
            if due_date:
                try:
                    date_obj = datetime.strptime(due_date, '%Y-%m-%d')
                    day_name = date_obj.strftime('%A')
                    due_datetime = f"{due_datetime} ({day_name})"
                except:
                    pass
                    
            # Shorten notes for display
            notes = row.get('Notes', '')
            if len(notes) > 50:
                notes = notes[:47] + "..."
                
            self.tree.insert("", "end", values=(
                index + 1,
                row.get('Project Name', ''),
                row.get('Status', ''),
                due_datetime,
                row.get('Progress', ''),
                reward_str,
                row.get('Project Link', ''),
                notes
            ))

    def update_dashboard(self):
        """Update dashboard statistics and charts"""
        # Update stats
        total_projects = len(self.df)
        active_projects = len(self.df[self.df['Status'] == 'Active']) if not self.df.empty else 0
        completed_projects = len(self.df[self.df['Status'] == 'Completed']) if not self.df.empty else 0
        total_value = self.df['Estimated Reward'].sum() if not self.df.empty else 0

        # Find the stats frame and update labels
        stats_frame = self.dashboard_tab.winfo_children()[0]
        for i, (value, color) in enumerate([
            (total_projects, "#4ECDC4"),
            (active_projects, "#45B7D1"),
            (completed_projects, "#96CEB4"),
            (f"Rp {total_value:,.2f}", "#FECA57")
        ]):
            stat_frame = stats_frame.grid_slaves(row=0, column=i)[0]
            # The value label is the second widget in the frame
            value_label = stat_frame.winfo_children()[1]
            value_label.configure(text=str(value))

        # Update charts - get the chart frames
        charts_frame = self.dashboard_tab.winfo_children()[1]
        status_frame = charts_frame.grid_slaves(row=0, column=0)[0]
        progress_frame = charts_frame.grid_slaves(row=0, column=1)[0]
        
        self.create_status_chart(status_frame)
        self.create_progress_chart(progress_frame)

    def check_reminders(self):
        """Check for due reminders, considering snoozed projects with thread safety"""
        if not self.reminder_active or self.df.empty or self.closing_app:
            return

        try:
            now = datetime.now()
            for index, project in self.df.iterrows():
                # Check if reminder is enabled for this project
                if not project.get('Reminder Enabled', True):
                    continue
                    
                # Create project identifier
                project_name = project.get('Project Name', '')
                due_date = project.get('Due Date', '')
                due_time = project.get('Due Time', '')
                project_id = f"{project_name}_{due_date}_{due_time}"
                
                # Check if project is snoozed
                if project_id in self.snoozed_projects:
                    if now < self.snoozed_projects[project_id]:
                        continue  # Still snoozed
                    else:
                        # Snooze time expired, remove from snoozed list
                        del self.snoozed_projects[project_id]
                        
                if project.get('Status') == 'Active' and project.get('Due Date') and project.get('Due Time'):
                    try:
                        due_date_str = project['Due Date']
                        due_time_str = project['Due Time']
                        
                        # Handle different date formats
                        if isinstance(due_date_str, datetime):
                            due_date = due_date_str
                        else:
                            due_date = datetime.strptime(due_date_str, "%Y-%m-%d")
                        
                        # Parse time
                        due_time = datetime.strptime(due_time_str, "%H:%M").time()
                        
                        # Combine date and time
                        due_datetime = datetime.combine(due_date, due_time)
                        
                        # Check if due time has passed and not yet notified
                        reminder_id = f"{project['Project Name']}_{project['Due Date']}_{project['Due Time']}"
                        
                        # Show reminder if due time has passed
                        if now >= due_datetime and reminder_id not in self.shown_reminders:
                            # Use thread-safe method to show reminder
                            if not self.closing_app and self.running:
                                self.show_reminder_thread_safe(project)
                                self.shown_reminders.add(reminder_id)
                                
                    except (ValueError, TypeError) as e:
                        print(f"Error parsing date/time for project {project.get('Project Name', 'Unknown')}: {e}")
                        continue
                        
        except Exception as e:
            print(f"Reminder check error: {e}")

    def show_reminder_thread_safe(self, project):
        """Thread-safe method to show reminder"""
        if self.running and not self.closing_app:
            try:
                # Schedule the reminder to be shown in the main thread
                self.after_idle(self._show_reminder_window, project)
            except (tk.TclError, RuntimeError):
                # Main thread might not be available
                print("Cannot show reminder: main thread not available")

    def show_reminder(self, project):
        """Show reminder notification as a popup window - legacy method"""
        self.show_reminder_thread_safe(project)

    def _show_reminder_window(self, project):
        """Actually show the reminder window (runs in main thread)"""
        if not self.running or self.closing_app or (self.reminder_window is not None and self.reminder_window.winfo_exists()):
            return
            
        try:
            # Create a top-level window for the reminder
            self.reminder_window = ctk.CTkToplevel(self)
            self.reminder_window.title("⏰ Airdrop Reminder")
            self.reminder_window.geometry("400x200")
            self.reminder_window.resizable(False, False)
            self.reminder_window.transient(self)
            self.reminder_window.lift()
            self.reminder_window.attributes('-topmost', True)
            
            # Center the reminder window
            self.reminder_window.update_idletasks()
            x = self.winfo_x() + (self.winfo_width() - self.reminder_window.winfo_width()) // 2
            y = self.winfo_y() + (self.winfo_height() - self.reminder_window.winfo_height()) // 2
            self.reminder_window.geometry(f"+{x}+{y}")
            
            # Reminder content
            title_label = ctk.CTkLabel(self.reminder_window, text="⏰ Airdrop Reminder", 
                                     font=ctk.CTkFont(size=16, weight="bold"))
            title_label.pack(pady=10)
            
            message = f"Project: {project.get('Project Name', 'Unknown')}\nDue: {project.get('Due Date', '')} {project.get('Due Time', '')}"
            message_label = ctk.CTkLabel(self.reminder_window, text=message, 
                                       font=ctk.CTkFont(size=14))
            message_label.pack(pady=10)
            
            # Button frame
            button_frame = ctk.CTkFrame(self.reminder_window)
            button_frame.pack(pady=20)
            
            # Snooze button
            snooze_btn = ctk.CTkButton(button_frame, text=f"Snooze ({self.snooze_var.get()}m)", 
                                     command=lambda: self.snooze_reminder(project))
            snooze_btn.pack(side="left", padx=10)
            
            # Dismiss button
            dismiss_btn = ctk.CTkButton(button_frame, text="Dismiss", 
                                      command=self.dismiss_reminder)
            dismiss_btn.pack(side="left", padx=10)
            
            # Play sound
            try:
                winsound.Beep(1000, 500)
            except:
                pass
            
            # Set focus to the reminder window
            self.reminder_window.focus_set()
            self.reminder_window.grab_set()
            
        except Exception as e:
            print(f"Error showing reminder window: {e}")

    def snooze_reminder(self, project):
        """Snooze the reminder for the specified duration"""
        if self.reminder_window:
            self.reminder_window.destroy()
            self.reminder_window = None
            
        # Calculate when to show the reminder again
        snooze_minutes = int(self.snooze_var.get())
        snooze_until = datetime.now() + timedelta(minutes=snooze_minutes)
        
        # Create a unique identifier for the project
        project_name = project.get('Project Name', '')
        due_date = project.get('Due Date', '')
        due_time = project.get('Due Time', '')
        project_id = f"{project_name}_{due_date}_{due_time}"
        
        # Store snooze time
        self.snoozed_projects[project_id] = snooze_until
        self.save_snooze_data()
        
        messagebox.showinfo("Snoozed", f"Reminder snoozed for {snooze_minutes} minutes.")

    def dismiss_reminder(self):
        """Dismiss the current reminder"""
        if self.reminder_window:
            self.reminder_window.destroy()
            self.reminder_window = None

    def save_data(self):
        """Save data to Excel with color coding"""
        try:
            # Use thread-safe lock for data access
            with self.data_lock:
                # Create a new workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Airdrop Data"
                
                # Write headers (exclude Reminder Enabled column)
                headers = [col for col in self.df.columns if col != 'Reminder Enabled']
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                
                # Define color fills for different statuses
                status_colors = {
                    'Active': PatternFill(start_color='4FC3F7', end_color='4FC3F7', fill_type='solid'),
                    'Completed': PatternFill(start_color='81C784', end_color='81C784', fill_type='solid'),
                    'Monitoring': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),
                    'Dropped': PatternFill(start_color='E57373', end_color='E57373', fill_type='solid')
                }
                
                # Write data with color coding (exclude Reminder Enabled column)
                for r_idx, row in self.df.iterrows():
                    for c_idx, col in enumerate(headers, 1):
                        value = row[col]
                        cell = ws.cell(row=r_idx+2, column=c_idx, value=value)
                        
                        # Apply color based on status
                        if col == 'Status' and value in status_colors:
                            cell.fill = status_colors[value]
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save the workbook
                wb.save(self.data_file_path)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save data: {str(e)}")

    def save_snooze_data(self):
        """Save snooze data to file"""
        try:
            snooze_data = {}
            for project_id, snooze_until in self.snoozed_projects.items():
                snooze_data[project_id] = snooze_until.isoformat()
            
            with open(self.snooze_file, 'w') as f:
                json.dump(snooze_data, f)
        except Exception as e:
            print(f"Error saving snooze data: {e}")

    def load_snooze_data(self):
        """Load snooze data from file"""
        try:
            if os.path.exists(self.snooze_file):
                with open(self.snooze_file, 'r') as f:
                    snooze_data = json.load(f)
                
                for project_id, snooze_until_str in snooze_data.items():
                    self.snoozed_projects[project_id] = datetime.fromisoformat(snooze_until_str)
        except Exception as e:
            print(f"Error loading snooze data: {e}")

    def load_data(self):
        """Load data from Excel with thread safety and error handling"""
        try:
            with self.data_lock:
                if os.path.exists(self.data_file_path):
                    df = pd.read_excel(self.data_file_path)
                    # Ensure all required columns exist
                    required_columns = ['Project Name', 'Status', 'Due Date', 'Due Time', 
                                      'Progress', 'Estimated Reward', 'Project Link', 'Notes', 'Reminder Enabled']
                    
                    for col in required_columns:
                        if col not in df.columns:
                            if col == 'Reminder Enabled':
                                df[col] = True
                            else:
                                df[col] = '' if col != 'Estimated Reward' else 0.0
                    
                    # Handle NaN values
                    df = df.fillna('')
                    
                    # Convert numpy.bool_ to Python bool for Reminder Enabled column
                    if 'Reminder Enabled' in df.columns:
                        df['Reminder Enabled'] = df['Reminder Enabled'].apply(lambda x: bool(x) if pd.notna(x) else True)
                    
                    return df
                else:
                    return pd.DataFrame(columns=[
                        'Project Name', 'Status', 'Due Date', 'Due Time',
                        'Progress', 'Estimated Reward', 'Project Link', 'Notes', 'Reminder Enabled'
                    ])
        except Exception as e:
            print(f"Load error: {e}")
            # Create backup of corrupted file
            if os.path.exists(self.data_file_path):
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                corrupted_backup = os.path.join(self.backup_dir, f"corrupted_backup_{timestamp}.xlsx")
                shutil.copy2(self.data_file_path, corrupted_backup)
                print(f"Created backup of corrupted file: {corrupted_backup}")
            
            messagebox.showwarning("Warning", "Data file is corrupted. Creating new dataset.")
            return pd.DataFrame(columns=[
                'Project Name', 'Status', 'Due Date', 'Due Time',
                'Progress', 'Estimated Reward', 'Project Link', 'Notes', 'Reminder Enabled'
            ])

    def create_backup(self):
        """Create data backup - including password file"""
        try:
            # First save the current data
            self.save_data()
            
            # Now create the backup
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(self.backup_dir, f"backup_{timestamp}.xlsx")
            shutil.copy2(self.data_file_path, backup_path)
            
            # Backup password file if it exists
            if os.path.exists(self.password_file):
                password_backup_path = os.path.join(self.security_backup_dir, f"password_backup_{timestamp}.txt")
                shutil.copy2(self.password_file, password_backup_path)
                
            # Backup snooze data
            if os.path.exists(self.snooze_file):
                snooze_backup_path = os.path.join(self.backup_dir, f"snooze_backup_{timestamp}.json")
                shutil.copy2(self.snooze_file, snooze_backup_path)
                
            messagebox.showinfo("Backup", f"Backup created successfully at:\n{backup_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Backup failed: {str(e)}")

    def export_excel(self):
        """Export data to Excel"""
        try:
            # Generate automatic filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"Airdrop_Data_Export_{timestamp}.xlsx"
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=default_filename
            )
            if file_path:
                # Save with color coding (exclude Reminder Enabled column)
                wb = Workbook()
                ws = wb.active
                ws.title = "Airdrop Data"
                
                # Write headers (exclude Reminder Enabled)
                headers = [col for col in self.df.columns if col != 'Reminder Enabled']
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                
                # Define color fills for different statuses
                status_colors = {
                    'Active': PatternFill(start_color='4FC3F7', end_color='4FC3F7', fill_type='solid'),
                    'Completed': PatternFill(start_color='81C784', end_color='81C784', fill_type='solid'),
                    'Monitoring': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),
                    'Dropped': PatternFill(start_color='E57373', end_color='E57373', fill_type='solid')
                }
                
                # Write data with color coding (exclude Reminder Enabled)
                for r_idx, row in self.df.iterrows():
                    for c_idx, col in enumerate(headers, 1):
                        value = row[col]
                        cell = ws.cell(row=r_idx+2, column=c_idx, value=value)
                        
                        # Apply color based on status
                        if col == 'Status' and value in status_colors:
                            cell.fill = status_colors[value]
                
                # Auto-adject column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(file_path)
                messagebox.showinfo("Success", "Data exported to Excel successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {str(e)}")

    def export_csv(self):
        """Export data to CSV"""
        try:
            # Generate automatic filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"Airdrop_Data_Export_{timestamp}.csv"
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialfile=default_filename
            )
            if file_path:
                # Exclude Reminder Enabled column
                export_df = self.df.drop(columns=['Reminder Enabled'], errors='ignore')
                export_df.to_csv(file_path, index=False)
                messagebox.showinfo("Success", "Data exported to CSV successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {str(e)}")

    def export_pdf(self):
        """Export data to PDF with proper formatting and word wrap in landscape mode"""
        try:
            # Generate automatic filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"Airdrop_Report_{timestamp}.pdf"
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                initialfile=default_filename
            )
            if file_path:
                # Create PDF document in landscape mode
                doc = SimpleDocTemplate(file_path, pagesize=landscape(letter))
                elements = []
                
                # Add title
                styles = getSampleStyleSheet()
                title = Paragraph("Airdrop Tracker Report", styles['Title'])
                elements.append(title)
                
                # Add date
                date_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                date = Paragraph(f"Generated on: {date_str}", styles['Normal'])
                elements.append(date)
                
                # Add some space
                elements.append(Paragraph("<br/><br/>", styles['Normal']))
                
                # Prepare data for table - exclude Reminder Enabled column
                columns_to_export = [
                    'Project Name', 'Status', 'Due DateTime', 'Progress', 
                    'Estimated Reward', 'Project Link', 'Notes'
                ]
                
                table_data = [columns_to_export]
                
                # Create a style for table cells with word wrap
                cell_style = styles['Normal']
                cell_style.wordWrap = 'CJK'
                
                for _, row in self.df.iterrows():
                    # Combine date and time for display
                    due_date = row.get('Due Date', '')
                    due_time = row.get('Due Time', '')
                    due_datetime = f"{due_date} {due_time}" if due_date and due_time else ''
                    
                    # Add day name if date is available
                    if due_date:
                        try:
                            date_obj = datetime.strptime(due_date, '%Y-%m-%d')
                            day_name = date_obj.strftime('%A')
                            due_datetime = f"{due_datetime} ({day_name})"
                        except:
                            pass
                    
                    table_data.append([
                        Paragraph(str(row.get('Project Name', '')), cell_style),
                        Paragraph(str(row.get('Status', '')), cell_style),
                        Paragraph(str(due_datetime), cell_style),
                        Paragraph(str(row.get('Progress', '')), cell_style),
                        Paragraph(f"Rp {row.get('Estimated Reward', 0):,.2f}", cell_style),
                        Paragraph(str(row.get('Project Link', '')), cell_style),
                        Paragraph(str(row.get('Notes', '')), cell_style)
                    ])
                
                # Create table with adjusted column widths for landscape
                col_widths = [
                    1.5*inch,  # Project Name
                    0.8*inch,  # Status
                    1.5*inch,  # Due DateTime
                    0.8*inch,  # Progress
                    1.0*inch,  # Estimated Reward
                    1.8*inch,  # Project Link
                    2.0*inch,  # Notes
                ]
                
                table = Table(table_data, colWidths=col_widths, repeatRows=1)
                
                # Add style to table
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(table)
                
                # Build PDF
                doc.build(elements)
                
                messagebox.showinfo("Success", "Data exported to PDF successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"PDF export failed: {str(e)}")

    def import_excel(self):
        """Import data from Excel file"""
        try:
            file_path = filedialog.askopenfilename(
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if file_path:
                new_df = pd.read_excel(file_path)
                
                # Check if the imported data has the required columns
                required_columns = ['Project Name', 'Status', 'Due Date', 'Due Time', 
                                  'Progress', 'Estimated Reward', 'Project Link', 'Notes', 'Reminder Enabled']
                
                for col in required_columns:
                    if col not in new_df.columns:
                        if col == 'Reminder Enabled':
                            new_df[col] = True
                        else:
                            new_df[col] = '' if col != 'Estimated Reward' else 0.0
                
                # Fix numpy.bool_ conversion issue
                if 'Reminder Enabled' in new_df.columns:
                    new_df['Reminder Enabled'] = new_df['Reminder Enabled'].apply(lambda x: bool(x) if pd.notna(x) else True)
                
                # Replace the current data with imported data
                with self.data_lock:
                    self.df = new_df
                self.save_data()
                self.update_table()
                self.update_dashboard()
                self.update_project_dropdown()
                messagebox.showinfo("Success", "Data imported from Excel successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Import failed: {str(e)}")

    def import_csv(self):
        """Import data from CSV file"""
        try:
            file_path = filedialog.askopenfilename(
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )
            if file_path:
                new_df = pd.read_csv(file_path)
                
                # Check if the imported data has the required columns
                required_columns = ['Project Name', 'Status', 'Due Date', 'Due Time', 
                                  'Progress', 'Estimated Reward', 'Project Link', 'Notes', 'Reminder Enabled']
                
                for col in required_columns:
                    if col not in new_df.columns:
                        if col == 'Reminder Enabled':
                            new_df[col] = True
                        else:
                            new_df[col] = '' if col != 'Estimated Reward' else 0.0
                
                # Replace the current data with imported data
                with self.data_lock:
                    self.df = new_df
                self.save_data()
                self.update_table()
                self.update_dashboard()
                self.update_project_dropdown()
                messagebox.showinfo("Success", "Data imported from CSV successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Import failed: {str(e)}")

    def toggle_theme(self):
        """Toggle between light and dark theme"""
        ctk.set_appearance_mode(self.theme_var.get())

    def toggle_reminders(self):
        """Toggle reminders on/off"""
        self.reminder_active = self.reminder_var.get()
        status = "enabled" if self.reminder_active else "disabled"
        messagebox.showinfo("Reminders", f"Reminders {status}")

    def toggle_auto_backup(self):
        """Toggle auto backup on/off"""
        self.auto_backup = self.backup_toggle.get()
        status = "enabled" if self.auto_backup else "disabled"
        messagebox.showinfo("Backup", f"Auto backup {status}")

    def update_project_dropdown(self):
        """Update the project dropdown in reminder settings"""
        if not self.df.empty:
            project_names = self.df['Project Name'].tolist()
            self.project_dropdown.configure(values=project_names)
            if project_names:
                self.project_var.set(project_names[0])
                self.on_project_select(project_names[0])
        else:
            self.project_dropdown.configure(values=[])
            self.project_var.set("")
            self.project_reminder_toggle.deselect()
            self.project_reminder_toggle.configure(state="disabled")

    def on_project_select(self, choice):
        """Handle project selection in reminder settings"""
        if not self.df.empty and choice:
            project_data = self.df[self.df['Project Name'] == choice].iloc[0]
            # Convert numpy.bool_ to Python bool if needed
            reminder_enabled = bool(project_data.get('Reminder Enabled', True))
            self.project_reminder_var.set(reminder_enabled)
            self.project_reminder_toggle.configure(state="normal")

    def toggle_project_reminder(self):
        """Toggle reminder for selected project"""
        selected_project = self.project_var.get()
        if selected_project and not self.df.empty:
            # Update the reminder setting for the selected project
            self.df.loc[self.df['Project Name'] == selected_project, 'Reminder Enabled'] = self.project_reminder_var.get()
            self.save_data()
            messagebox.showinfo("Success", f"Reminder setting updated for {selected_project}")

    def test_reminder(self):
        """Test reminder notification"""
        self.show_reminder({
            'Project Name': 'Test Project',
            'Due Date': datetime.now().strftime('%Y-%m-%d'),
            'Due Time': datetime.now().strftime('%H:%M')
        })

    def delete_all_data(self):
        """Delete all data with confirmation"""
        if messagebox.askyesno("Confirm Deletion", "Are you sure you want to delete ALL data? This action cannot be undone!"):
            try:
                with self.data_lock:
                    self.df = pd.DataFrame(columns=[
                        'Project Name', 'Status', 'Due Date', 'Due Time',
                        'Progress', 'Estimated Reward', 'Project Link', 'Notes', 'Reminder Enabled'
                    ])
                self.shown_reminders.clear()
                self.snoozed_projects.clear()
                self.save_data()
                self.update_table()
                self.update_dashboard()
                self.update_project_dropdown()
                messagebox.showinfo("Success", "All data has been deleted successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete data: {str(e)}")

    def show_dashboard(self):
        """Show dashboard tab"""
        self.tabview.set("📊 Dashboard")
        self.update_dashboard()

    def show_data_manager(self):
        """Show data manager tab"""
        self.tabview.set("📋 Data")

    def show_reminder_settings(self):
        """Show reminder settings tab"""
        self.tabview.set("⏰ Reminders")
        self.update_project_dropdown()

    def show_settings(self):
        """Show settings tab"""
        self.tabview.set("⚙️ Settings")

    def show_about(self):
        """Show about tab"""
        self.tabview.set("ℹ️ About")

    def on_closing(self):
        """Handle application closing - minimize to tray instead of closing"""
        if not self.closing_app:
            # If tray is available, minimize to tray
            if self.tray_icon:
                self.minimize_to_tray()
                return
            else:
                # Ask user what to do if tray is not available
                result = messagebox.askyesnocancel(
                    "Close Application",
                    "System tray is not available.\n\nDo you want to:\n• Yes: Close completely\n• No: Minimize to taskbar\n• Cancel: Continue running"
                )
                if result is True:
                    # Close completely
                    self.closing_app = True
                    self.quit_application()
                elif result is False:
                    # Minimize to taskbar
                    self.iconify()
                # If Cancel (None), do nothing
        else:
            # Force quit
            self.quit_application()

    def protocol_handler(self, protocol, func):
        """Custom protocol handler to manage window close behavior"""
        if protocol == "WM_DELETE_WINDOW":
            super().protocol(protocol, self.on_closing)
        else:
            super().protocol(protocol, func)


# Run the application
if __name__ == "__main__":
    try:
        app = PremiumAirdropTracker()
        app.protocol("WM_DELETE_WINDOW", app.on_closing)
        app.mainloop()
    except Exception as e:
        print(f"Application error: {e}")
        import traceback
        traceback.print_exc()
        os._exit(1)